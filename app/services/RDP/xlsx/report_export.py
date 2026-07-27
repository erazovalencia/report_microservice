from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from typing import List
from datetime import datetime, timezone, timedelta

BOGOTA_TZ = timezone(timedelta(hours=-5))

from ...base import BaseExportService
from ....models.RdpModel import RdpExportRow

COLORS = {
    "header_bg":       "1F3864",
    "header_fg":       "FFFFFF",
    "turno_bg":        "D9E8D9",
    "ausencia_bg":     "FDEBD0",
    "hora_extra_bg":   "D6EAF8",
    "row_even":        "F5F5F5",
    "row_odd":         "FFFFFF",
    "subheader_bg":    "BDD7EE",
    "approved_fg":     "1E8449",
    "pending_fg":      "B7950B",
    "correction_fg":   "C0392B",
}

ESTADO_LABEL = {
    "submitted":   "En revisión N1",
    "approved_l1": "En revisión N2",
    "approved":    "Aprobado",
    "draft":       "Borrador",
    "correction":  "En corrección",
}

TIPO_LABEL = {
    "TURNO":      "Turno",
    "AUSENTISMO": "Ausentismo",
    "HORA_EXTRA": "Hora Extra",
}

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# Estilos de celda de datos reutilizados por referencia — antes se instanciaba un
# Font/Alignment nuevo por CADA celda (27 cols x hasta 30k filas = hasta ~810k objetos),
# que es el patrón conocido más lento en openpyxl (cada estilo nuevo se registra y
# deduplica contra la StyleArray global del workbook).
DATA_FONT    = Font(size=9)
ALIGN_CENTER = Alignment(vertical="center", horizontal="center")
ALIGN_LEFT   = Alignment(vertical="center", horizontal="left")
CENTERED_COLS = {1, 3, 5, 6, 10, 11, 12, 13, 14, 15}

_ROW_BG_TIPO = {
    "TURNO":      (COLORS["turno_bg"],      "EAF4EA"),
    "AUSENTISMO": (COLORS["ausencia_bg"],   "FEF5E7"),
    "HORA_EXTRA": (COLORS["hora_extra_bg"], "EBF5FB"),
    None:         (COLORS["row_even"],      COLORS["row_odd"]),
}

# PatternFill también reutilizado — antes uno nuevo por fila (hasta 30k); el color
# real solo depende de (tipo, paridad de la fila), así que son 8 combinaciones fijas.
ROW_FILLS = {
    (tipo, parity): PatternFill(start_color=colors[parity], fill_type="solid")
    for tipo, colors in _ROW_BG_TIPO.items()
    for parity in (0, 1)
}

HEADERS = [
    ("Fecha",                  16),
    ("Unidad Organizativa",    28),
    ("Cédula",                 14),
    ("Nombre",                 34),
    ("Tipo",                   14),
    ("Turno / Ausencia",       18),
    ("Tipo de Bono",           16),
    ("Centro de Costo",        18),
    ("Pozo / Ubicación",        22),
    ("H. Entrada",             12),
    ("H. Salida",              12),
    ("Total Horas",            12),
    ("Rotación",               12),
    ("Entrada BioStar",        16),
    ("Salida BioStar",         16),
    ("HE Total",               10),
    ("HE Diurna",              10),
    ("HE Diurna Fest.",        13),
    ("HE Nocturna",            11),
    ("HE Nocturna Fest.",      14),
    ("HRD",                    10),
    ("RNF",                    10),
    ("Impacto Vac.",           13),
    ("Impacto Comp.",          13),
    ("Estado Reporte",         18),
    ("Notas Turno",            25),
    ("Notas Ausencia",         25),
]


def _fmt_biostar(iso: str) -> str:
    if not iso:
        return ""
    try:
        dt = datetime.fromisoformat(iso.replace("Z", "+00:00"))
        return dt.astimezone(BOGOTA_TZ).strftime("%H:%M")
    except Exception:
        return iso


def _row_fill(tipo: str, row_index: int) -> PatternFill:
    tipo_key = tipo if tipo in ("TURNO", "AUSENTISMO", "HORA_EXTRA") else None
    return ROW_FILLS[(tipo_key, row_index % 2)]


class RdpReportExportService(BaseExportService):

    def generate_file(self, data: List[RdpExportRow], options=None) -> io.BytesIO:
        wb = Workbook()
        self._build_report_sheet(wb, data)
        self._build_legend_sheet(wb)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def _build_report_sheet(self, wb: Workbook, rows: List[RdpExportRow]):
        ws = wb.active
        ws.title = "Reporte RDP"
        ws.freeze_panes = "A3"

        # Title row
        ws.merge_cells(f"A1:{get_column_letter(len(HEADERS))}1")
        title_cell = ws["A1"]
        title_cell.value = "REPORTE RDP — Erazo Valencia"
        title_cell.font = Font(bold=True, size=13, color=COLORS["header_fg"])
        title_cell.fill = PatternFill(start_color=COLORS["header_bg"], fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        # Header row
        for col, (label, width) in enumerate(HEADERS, start=1):
            cell = ws.cell(row=2, column=col, value=label)
            cell.font = Font(bold=True, size=10, color=COLORS["header_fg"])
            cell.fill = PatternFill(start_color=COLORS["header_bg"], fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[2].height = 28

        # Data rows
        for i, row in enumerate(rows):
            r = i + 3
            fill = _row_fill(row.tipo, i)

            values = [
                row.fecha,
                row.unidadOrganizativa,
                row.identificacion,
                row.nombre,
                TIPO_LABEL.get(row.tipo, row.tipo),
                row.turnoAusencia,
                row.tipoBono or "",
                row.centroCosto or "",
                row.actividad or "",
                row.horarioEntrada or "",
                row.horarioSalida or "",
                row.totalHoras or "",
                row.rotacion or "",
                _fmt_biostar(row.biostarEntrada),
                _fmt_biostar(row.biostarSalida),
                row.heTotal if row.heTotal else "",
                row.heDiurna if row.heDiurna else "",
                row.heDiurnaFestiva if row.heDiurnaFestiva else "",
                row.heNocturna if row.heNocturna else "",
                row.heNocturnaFestiva if row.heNocturnaFestiva else "",
                row.hrd if row.hrd else "",
                row.rnf if row.rnf else "",
                round(row.impactoVacaciones, 4) if row.impactoVacaciones else "",
                round(row.impactoCompensatorios, 4) if row.impactoCompensatorios else "",
                ESTADO_LABEL.get(row.estadoReporte, row.estadoReporte),
                row.notasTurno    or "",
                row.notasAusencia or "",
            ]

            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.fill = fill
                cell.border = BORDER
                # cols centradas: Fecha(1), Cédula(3), Tipo(5), Turno/Aus(6), H.Entrada(10), H.Salida(11), TotalHoras(12), Rotación(13), BioStar(14,15)
                cell.alignment = ALIGN_CENTER if col in CENTERED_COLS else ALIGN_LEFT
                cell.font = DATA_FONT

            ws.row_dimensions[r].height = 16

        # Auto-filter on header row
        ws.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS))}{len(rows) + 2}"

    def _build_legend_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Leyenda")

        legend = [
            ("TURNO",      "Verde",   "Día de trabajo según turno asignado"),
            ("AUSENTISMO", "Naranja", "Día de ausencia (vacaciones, incapacidad, licencia, etc.)"),
            ("HORA_EXTRA", "Azul",    "Horas extras trabajadas fuera del turno"),
        ]

        ws["A1"] = "Tipo de Registro"
        ws["B1"] = "Color"
        ws["C1"] = "Descripción"
        for col in ["A", "B", "C"]:
            cell = ws[f"{col}1"]
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=COLORS["header_bg"], fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        fills = {"TURNO": "D9E8D9", "AUSENTISMO": "FDEBD0", "HORA_EXTRA": "D6EAF8"}
        for i, (tipo, color, desc) in enumerate(legend, start=2):
            ws.cell(row=i, column=1, value=tipo).fill = PatternFill(start_color=fills[tipo], fill_type="solid")
            ws.cell(row=i, column=2, value=color)
            ws.cell(row=i, column=3, value=desc)

        ws["A6"] = "Impacto Vacaciones"
        ws["B6"] = "-0.0417 por día de ausencia en tipos 300/500/601/602"
        ws["A7"] = "Impacto Compensatorios"
        ws["B7"] = "Fracción acumulada según rotación (3x1=0.33, 2x1=0.50, 1x1=1.00, 5x2=0.40, 6x1=0.16)"
        ws["A8"] = "Entrada/Salida BioStar"
        ws["B8"] = "Primer y último acceso registrado en BioStar ese día (puede estar vacío si no hay datos)"

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 55

    def get_content_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def get_file_extension(self) -> str:
        return ".xlsx"
