from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import io
from typing import List

from ...base import BaseExportService
from ....models.RdpModel import RdpCatalogEntry

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _fmt_hour(value) -> str:
    """Convierte una hora decimal del catálogo (6.0, 8.5) a 'HH:MM'. None/'' → ''."""
    if value is None or value == "":
        return ""
    try:
        h = float(value)
    except (ValueError, TypeError):
        return ""
    hh = int(h)
    mm = int(round((h - hh) * 60))
    if mm == 60:
        hh += 1
        mm = 0
    return f"{hh:02d}:{mm:02d}"

HEADER_BG   = "1F3864"
HEADER_FG   = "FFFFFF"
REQUIRED_BG = "FFF3CD"
OPTIONAL_BG = "F0F0F0"

# Sin fila de ejemplo en la hoja de datos: una cédula "de muestra" se confunde con
# un empleado real y el parser tendría que adivinar dónde empiezan los datos.
# Los ejemplos viven en los hints (fila 4) y en las hojas de catálogo.
COLUMNS = [
    ("Identificación",   "Cédula del empleado",                                True,  16),
    ("Turno",            "Código de turno (ver hoja Turnos)",                  False, 14),
    ("Tipo Ausencia",    "Nombre de ausencia (ver hoja Ausencias)",            False, 22),
    ("Tipo de Bono",     "Código de bono (ver hoja Bonos)",                    False, 16),
    ("Centro de Costo",  "Código centro de costo (ver hoja Costos)",           False, 18),
    ("Pozo / Ubicacion", "Descripción de actividad del turno",                 False, 22),
    ("Hora Ingreso",     "Hora entrada real HH:MM (ej. 06:00)",                False, 14),
    ("Hora Salida",      "Hora salida real HH:MM  (ej. 18:00)",                False, 14),
    ("Salida Día Siguiente", "Sí si la salida fue al día siguiente (turno >24h)", False, 18),
    ("Notas",            "Observaciones del registro",                         False, 28),
]


class RdpImportTemplateService(BaseExportService):

    def generate_file(
        self,
        data,
        options=None,
    ) -> io.BytesIO:
        shifts      = data.get("shifts",      [])
        absences    = data.get("absences",    [])
        bonuses     = data.get("bonuses",     [])
        workCenters = data.get("workCenters", [])
        employees   = data.get("employees",   [])

        wb = Workbook()
        self._build_main_sheet(wb, employees)
        self._build_shifts_sheet(wb, shifts)
        self._build_catalog_sheet(wb, "Ausencias", absences,    ["Nombre", "Código"],  display="name")
        self._build_catalog_sheet(wb, "Bonos",     bonuses,     ["Código", "Nombre"],  display="code")
        self._build_catalog_sheet(wb, "Costos",    workCenters, ["Código", "Nombre"],  display="name")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    @staticmethod
    def _theoretical_formula(row: int, col_index: int) -> str:
        """VLOOKUP del código de turno (col B de la fila) en la hoja 'Turnos' → hora teórica.
        col_index: 3 = Ingreso Teórico, 4 = Salida Teórica."""
        return f'=IFERROR(VLOOKUP($B{row},Turnos!$A:$D,{col_index},FALSE),"")'

    def _build_main_sheet(self, wb: Workbook, employees: list):
        ws = wb.active
        ws.title = "Reporte"
        ws.freeze_panes = "A4"

        # Columnas extra al final (referencia, no importadas):
        #   Nombre · Ingreso Teórico · Salida Teórica
        ref_col    = len(COLUMNS) + 1   # J — Nombre
        ing_col    = len(COLUMNS) + 2   # K — Ingreso Teórico (VLOOKUP al turno)
        sal_col    = len(COLUMNS) + 3   # L — Salida Teórica  (VLOOKUP al turno)
        total_cols = len(COLUMNS) + 3

        # Title
        ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
        t = ws["A1"]
        t.value = "PLANTILLA IMPORTACIÓN MASIVA RDP — Erazo Valencia"
        t.font = Font(bold=True, size=12, color=HEADER_FG)
        t.fill = PatternFill(start_color=HEADER_BG, fill_type="solid")
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        # Subtitle / instructions
        ws.merge_cells(f"A2:{get_column_letter(total_cols)}2")
        s = ws["A2"]
        s.value = (
            "Instrucciones: complete una fila por empleado. "
            "Columnas con * son obligatorias. "
            "Ingrese Turno O Tipo Ausencia (no ambos). "
            "Use los códigos de las hojas de catálogo. "
            "Las columnas Nombre / Ingreso Teórico / Salida Teórica son de referencia (no se importan)."
        )
        s.font = Font(italic=True, size=9, color="555555")
        s.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 18

        # Column headers (row 3) — data columns
        for col, (label, hint, required, width) in enumerate(COLUMNS, start=1):
            header_label = f"{label} *" if required else label
            cell = ws.cell(row=3, column=col, value=header_label)
            cell.font = Font(bold=True, size=10, color=HEADER_FG)
            cell.fill = PatternFill(start_color=HEADER_BG, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[3].height = 24

        # Reference column headers (Nombre · Ingreso Teórico · Salida Teórica)
        ref_headers = [
            (ref_col, "Nombre (referencia)", 30),
            (ing_col, "Ingreso Teórico",     16),
            (sal_col, "Salida Teórica",      16),
        ]
        for rc, label, width in ref_headers:
            ref_header = ws.cell(row=3, column=rc, value=label)
            ref_header.font = Font(bold=True, size=9, color="777777")
            ref_header.fill = PatternFill(start_color="E8E8E8", fill_type="solid")
            ref_header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ref_header.border = BORDER
            ws.column_dimensions[get_column_letter(rc)].width = width

        # Hint row (row 4)
        for col, (_, hint, required, _) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=4, column=col, value=hint)
            cell.font = Font(italic=True, size=8, color="666666")
            cell.fill = PatternFill(
                start_color=REQUIRED_BG if required else OPTIONAL_BG,
                fill_type="solid",
            )
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER
        for rc, hint_txt in (
            (ref_col, "Solo referencia — no se importa"),
            (ing_col, "Automático según turno"),
            (sal_col, "Automático según turno"),
        ):
            ws.cell(row=4, column=rc, value=hint_txt).font = Font(italic=True, size=8, color="999999")
        ws.row_dimensions[4].height = 20

        # Data rows always start at row 5 — the parser relies on this fixed layout.
        start_data_row = 5

        # Pre-fill employee rows or blank rows
        rows_to_fill = employees
        num_rows     = max(len(rows_to_fill), 50)

        for i in range(num_rows):
            r  = start_data_row + i
            bg = "F0F8FF" if i % 2 == 0 else "FFFFFF"

            emp    = rows_to_fill[i] if i < len(rows_to_fill) else None
            cedula = emp.get("cedula", "") if emp else ""
            nombre = emp.get("nombre", "") if emp else ""

            for col in range(1, total_cols + 1):
                cell = ws.cell(row=r, column=col)
                cell.border = BORDER

                if col == 1:
                    # Cédula pre-rellena — fondo destacado si tiene datos
                    cell.value = cedula
                    cell.font = Font(size=9, bold=bool(cedula))
                    cell.fill = PatternFill(
                        start_color="D9EAD3" if cedula else bg,
                        fill_type="solid",
                    )
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col == ref_col:
                    # Nombre referencia
                    cell.value = nombre
                    cell.font = Font(size=9, color="666666", italic=True)
                    cell.fill = PatternFill(start_color="F5F5F5", fill_type="solid")
                    cell.alignment = Alignment(vertical="center")
                elif col in (ing_col, sal_col):
                    # Horas teóricas — VLOOKUP al turno seleccionado en col B (solo referencia)
                    lookup_idx = 3 if col == ing_col else 4
                    cell.value = self._theoretical_formula(r, lookup_idx)
                    cell.font = Font(size=9, color="1F3864")
                    cell.fill = PatternFill(start_color="F5F5F5", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.value = ""
                    cell.fill = PatternFill(start_color=bg, fill_type="solid")

            ws.row_dimensions[r].height = 16

        # Data validation — dropdowns por columna
        self._add_validations(ws, start_data_row, start_data_row + num_rows - 1)

    def _add_validations(self, ws, first_row: int, last_row: int):
        # B — Turno
        dv_turno = DataValidation(
            type="list",
            formula1="Turnos!$A$2:$A$200",
            showDropDown=False,
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Código inválido",
            error="Seleccione un turno de la hoja 'Turnos'",
        )
        dv_turno.sqref = f"B{first_row}:B{last_row}"
        ws.add_data_validation(dv_turno)

        # C — Tipo Ausencia
        dv_ausencia = DataValidation(
            type="list",
            formula1="Ausencias!$A$2:$A$200",
            showDropDown=False,
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Código inválido",
            error="Seleccione un código de la hoja 'Ausencias'",
        )
        dv_ausencia.sqref = f"C{first_row}:C{last_row}"
        ws.add_data_validation(dv_ausencia)

        # D — Tipo de Bono
        dv_bono = DataValidation(
            type="list",
            formula1="Bonos!$A$2:$A$200",
            showDropDown=False,
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Código inválido",
            error="Seleccione un bono de la hoja 'Bonos'",
        )
        dv_bono.sqref = f"D{first_row}:D{last_row}"
        ws.add_data_validation(dv_bono)

        # E — Centro de Costo
        dv_costo = DataValidation(
            type="list",
            formula1="Costos!$A$2:$A$200",
            showDropDown=False,
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Código inválido",
            error="Seleccione un centro de costo de la hoja 'Costos'",
        )
        dv_costo.sqref = f"E{first_row}:E{last_row}"
        ws.add_data_validation(dv_costo)

        # I — Salida Día Siguiente (Sí/No)
        dv_next_day = DataValidation(
            type="list",
            formula1='"Sí,No"',
            showDropDown=False,
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Valor inválido",
            error="Seleccione Sí o No",
        )
        dv_next_day.sqref = f"I{first_row}:I{last_row}"
        ws.add_data_validation(dv_next_day)

    def _build_shifts_sheet(self, wb: Workbook, shifts: list):
        """
        Hoja 'Turnos' — col A=Código (al dropdown), B=Nombre, C=Ingreso Teórico, D=Salida Teórica.
        Las columnas C/D alimentan el VLOOKUP de la hoja 'Reporte' (solo referencia, no se importan).
        """
        ws = wb.create_sheet("Turnos")
        ws.protection.sheet = True
        ws.protection.selectLockedCells = False
        ws.protection.selectUnlockedCells = False

        headers = ["Código", "Nombre", "Ingreso Teórico", "Salida Teórica"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color=HEADER_FG)
            cell.fill = PatternFill(start_color=HEADER_BG, fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER

        if not shifts:
            ws.cell(row=2, column=1, value="(sin registros)")
        else:
            for i, entry in enumerate(shifts, start=2):
                get = (lambda k: entry.get(k, "")) if isinstance(entry, dict) else (lambda k: getattr(entry, k, ""))
                code  = get("code")
                name  = get("name")
                start = _fmt_hour(get("startHour"))
                end   = _fmt_hour(get("endHour"))
                bg    = "F5F5F5" if i % 2 == 0 else "FFFFFF"
                fill  = PatternFill(start_color=bg, fill_type="solid")

                values = [(code, True, "000000"), (name, False, "888888"),
                          (start, False, "1F3864"), (end, False, "1F3864")]
                for col, (val, bold, color) in enumerate(values, start=1):
                    c = ws.cell(row=i, column=col, value=val)
                    c.fill = fill
                    c.border = BORDER
                    c.font = Font(bold=bold, size=9, color=color)
                    c.alignment = Alignment(horizontal="center" if col >= 3 else "left")

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 16

    def _build_catalog_sheet(
        self,
        wb: Workbook,
        title: str,
        entries: list,
        col_headers: list,
        display: str = "code",  # "code" | "code_name" | "name"
    ):
        ws = wb.create_sheet(title)

        # Proteger la hoja para que el reportador no la modifique accidentalmente
        ws.protection.sheet = True
        ws.protection.selectLockedCells = False    # puede seleccionar para ver
        ws.protection.selectUnlockedCells = False

        for col, header in enumerate(col_headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color=HEADER_FG)
            cell.fill = PatternFill(start_color=HEADER_BG, fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER

        if not entries:
            ws.cell(row=2, column=1, value="(sin registros)")
            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 40
            return

        max_a = 10
        max_b = 10
        for i, entry in enumerate(entries, start=2):
            code = entry.get("code", "") if isinstance(entry, dict) else getattr(entry, "code", "")
            name = entry.get("name", "") if isinstance(entry, dict) else getattr(entry, "name", "")
            bg   = "F5F5F5" if i % 2 == 0 else "FFFFFF"
            fill = PatternFill(start_color=bg, fill_type="solid")

            if display == "code_name":
                # Col A = "código — nombre"  (el que va al dropdown)
                # Col B = código limpio       (referencia visual)
                display_val = f"{code} — {name}" if name else code
                c1 = ws.cell(row=i, column=1, value=display_val)
                c2 = ws.cell(row=i, column=2, value=code)
                max_a = max(max_a, len(display_val))
                max_b = max(max_b, len(code))
            elif display == "name":
                # Col A = nombre (al dropdown)
                # Col B = código (referencia)
                c1 = ws.cell(row=i, column=1, value=name)
                c2 = ws.cell(row=i, column=2, value=code)
                max_a = max(max_a, len(name))
                max_b = max(max_b, len(code))
            else:
                # default: col A = código, col B = nombre
                c1 = ws.cell(row=i, column=1, value=code)
                c2 = ws.cell(row=i, column=2, value=name)
                max_a = max(max_a, len(code))
                max_b = max(max_b, len(name))

            c1.fill = fill; c1.border = BORDER; c1.font = Font(bold=(display == "code"), size=9)
            c2.fill = fill; c2.border = BORDER; c2.font = Font(size=9, color="888888")

        ws.column_dimensions["A"].width = min(max_a + 4, 55)
        ws.column_dimensions["B"].width = min(max_b + 4, 20)

    def get_content_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def get_file_extension(self) -> str:
        return ".xlsx"
