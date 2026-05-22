from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import io

from ...base import BaseExportService
from ....models.RdpModel import RdpConsolidatedRequest

NAVY   = "1F3864"
WHITE  = "FFFFFF"
HEADER_ROW_BG = "BDD7EE"
ALT_ROW        = "EBF3FB"

thin   = Side(style="thin",   color="AAAAAA")
medium = Side(style="medium",  color="1F3864")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
BORDER_HEADER = Border(left=medium, right=medium, top=medium, bottom=medium)

# Columnas de la tabla de datos
COLS = [
    ("FECHA",           10),
    ("TURNO",           8),
    ("DESCRIPCIÓN",     28),
    ("HORA DE\nINGRESO", 9),
    ("HORA DE\nSALIDA",  9),
    ("TOTAL\nHORAS",     8),
    ("HED\nADICIONAL",  8),
    ("HEN\nADICIONAL",  8),
    ("HEDF\nADICIONAL", 8),
    ("HENF\nADICIONAL", 8),
    ("CENTRO DE\nCOSTOS", 16),
    ("POZO /\nUBICACIÓN",  28),
    ("NOTAS",              30),
]
N_COLS = len(COLS)


def _cell(ws, row, col, value="", bold=False, size=9, color="000000",
          bg=None, align_h="center", align_v="center", wrap=False, border=BORDER):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=size, color=color)
    c.alignment = Alignment(horizontal=align_h, vertical=align_v, wrap_text=wrap)
    c.border = border
    if bg:
        c.fill = PatternFill(start_color=bg, fill_type="solid")
    return c


class RdpConsolidatedService(BaseExportService):

    def generate_file(self, data: RdpConsolidatedRequest, options=None) -> io.BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Consolidado"

        # Ancho de columnas
        for i, (_, w) in enumerate(COLS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        row = 1

        # ── Título ────────────────────────────────────────────────────────────
        ws.merge_cells(f"A{row}:{get_column_letter(N_COLS)}{row}")
        per = data.periodo
        title = (
            f"REPORTE CONSOLIDADO DE TURNOS PARA EL PERIODO "
            f"COMPRENDIDO ENTRE {per.get('from','')} - {per.get('to','')}"
        )
        c = ws.cell(row=row, column=1, value=title)
        c.font = Font(bold=True, size=11, color=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 22
        row += 1

        # ── Espacio ───────────────────────────────────────────────────────────
        ws.row_dimensions[row].height = 8
        row += 1

        # ── Datos del empleado ────────────────────────────────────────────────
        emp = data.empleado

        def emp_row(label1, val1, label2="", val2="", label3="", val3=""):
            nonlocal row
            # Fila con hasta 3 pares label/valor distribuidos en las columnas
            ws.row_dimensions[row].height = 14
            _cell(ws, row, 1, label1, bold=True, align_h="left", border=Border())
            ws.merge_cells(f"B{row}:D{row}")
            _cell(ws, row, 2, val1, align_h="left", border=Border())
            if label2:
                _cell(ws, row, 5, label2, bold=True, align_h="left", border=Border())
                ws.merge_cells(f"F{row}:G{row}")
                _cell(ws, row, 6, val2, align_h="left", border=Border())
            if label3:
                _cell(ws, row, 8, label3, bold=True, align_h="left", border=Border())
                ws.merge_cells(f"I{row}:{get_column_letter(N_COLS)}{row}")
                _cell(ws, row, 9, val3, align_h="left", border=Border())
            row += 1

        emp_row("Identificación:", emp.identificacion, "Nombre:", emp.nombre, "Edad:", str(emp.edad) if emp.edad else "")
        emp_row("Cargo:",          emp.cargo,         "Unidad Organizativa:", emp.unidadOrganizativa, "Sexo:", emp.sexo)
        emp_row("Tipo de nómina:", emp.tipoNomina,    "Empresa:", emp.empresa)

        row += 1  # espacio

        # ── Encabezado de tabla ───────────────────────────────────────────────
        ws.row_dimensions[row].height = 30
        for col_i, (label, _) in enumerate(COLS, start=1):
            _cell(ws, row, col_i, label, bold=True, size=8, color=WHITE,
                  bg=NAVY, wrap=True, border=BORDER_HEADER)
        row += 1

        # ── Filas de datos ────────────────────────────────────────────────────
        for i, fila in enumerate(data.filas):
            bg = ALT_ROW if i % 2 == 0 else WHITE
            ws.row_dimensions[row].height = 14
            vals = [
                fila.fecha,
                fila.turno,
                fila.descripcion,
                fila.horaIngreso,
                fila.horaSalida,
                fila.totalHoras   if fila.totalHoras  else "",
                fila.hed          if fila.hed          else "",
                fila.hen          if fila.hen          else "",
                fila.hedf         if fila.hedf         else "",
                fila.henf         if fila.henf         else "",
                fila.centroCostos,
                fila.actividad,
                fila.notas,
            ]
            for col_i, val in enumerate(vals, start=1):
                align = "center" if col_i in (1, 2, 4, 5, 6, 7, 8, 9, 10) else "left"
                _cell(ws, row, col_i, val, size=9, bg=bg, align_h=align)
            row += 1

        # ── Fila vacía separadora ─────────────────────────────────────────────
        ws.row_dimensions[row].height = 8
        row += 1

        # ── Pie: firma + totales HE segregados ────────────────────────────────
        # Solo muestra un tipo si su total > 0; siempre muestra el total general.
        he_totals = [
            ("HED ADICIONAL",  data.totalHed,  "E8F5E9"),
            ("HEN ADICIONAL",  data.totalHen,  "EDE7F6"),
            ("HEDF ADICIONAL", data.totalHedf, "FFF3E0"),
            ("HENF ADICIONAL", data.totalHenf, "F3E5F5"),
        ]
        visible_totals = [(lbl, val, bg) for lbl, val, bg in he_totals if val > 0]
        visible_totals.append(("TOTAL HE", data.totalHorasAdicionales, "FFF2CC"))

        for i, (lbl, val, bg_color) in enumerate(visible_totals):
            ws.row_dimensions[row].height = 16
            if i == 0:
                ws.merge_cells(f"A{row}:D{row}")
                _cell(ws, row, 1, "FIRMA________________________",
                      bold=False, size=9, align_h="left", border=Border())
            ws.merge_cells(f"F{row}:I{row}")
            _cell(ws, row, 6, lbl, bold=True, size=9,
                  bg=HEADER_ROW_BG, border=BORDER)
            ws.merge_cells(f"J{row}:{get_column_letter(N_COLS)}{row}")
            _cell(ws, row, 10, val if val else 0, bold=True, size=10,
                  bg=bg_color, border=BORDER)
            row += 1
        row -= 1  # el freeze_panes usa el row final

        # Freeze panes debajo del encabezado (row 7 = primera fila de datos aprox)
        ws.freeze_panes = ws.cell(row=8, column=1)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def get_content_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def get_file_extension(self) -> str:
        return ".xlsx"
