import io
import openpyxl
from typing import List, Dict, Any, Optional

# Columnas esperadas en la plantilla (índice 0-based)
COL_IDENTIFICACION  = 0
COL_TURNO           = 1
COL_TIPO_AUSENCIA   = 2
COL_TIPO_BONO       = 3
COL_PROYECTO        = 4
COL_ACTIVIDAD       = 5
COL_HORA_INGRESO    = 6
COL_HORA_SALIDA     = 7
COL_SALIDA_DIA_SIG  = 8
COL_NOTAS           = 9

# Encabezado fijo de la plantilla: título + instrucción + header + hint → datos desde fila 5.
SKIP_ROWS = 4

# Plantillas legacy (grupo sin empleados) traían una fila de ejemplo en la fila 5
# con esta cédula ficticia. Se salta solo si aparece intacta.
LEGACY_EXAMPLE_CEDULA = "1069742877"


def _str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _extract_code(val) -> str:
    """Extrae el código de un valor que puede ser 'código', 'código — nombre', o 'nombre'."""
    s = _str(val)
    if " — " in s:
        return s.split(" — ")[0].strip()
    return s


def _float_field(val) -> Optional[float]:
    if val is None or _str(val) == "":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _bool_field(val) -> bool:
    """Interpreta Sí/No (también si, true, 1, x, yes) como booleano. Vacío → False."""
    s = _str(val).lower()
    return s in ("sí", "si", "true", "1", "x", "yes", "verdadero")


def _time_field(val) -> Optional[float]:
    """Acepta HH:MM (ej. '06:00') o decimal (ej. 6.5). Devuelve float de hora."""
    s = _str(val)
    if not s:
        return None
    if ":" in s:
        parts = s.split(":")
        try:
            h = int(parts[0])
            m = int(parts[1]) if len(parts) > 1 else 0
            return h + m / 60.0
        except (ValueError, IndexError):
            return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _detect_skip_rows(ws) -> int:
    """
    Los datos siempre empiezan en la fila 5. Única excepción: plantillas legacy
    cuya fila 5 es la fila de ejemplo intacta (cédula ficticia) — esa se salta.
    Una fila 5 con datos pero sin cédula ya NO se descarta en silencio: se parsea
    y sale con error visible "Identificación es obligatoria".
    """
    cell_a5 = ws.cell(row=5, column=1).value
    ref_a5  = _str(ws.cell(row=5, column=11).value).lower()  # col K: "Nombre (referencia)"
    if _str(cell_a5) == LEGACY_EXAMPLE_CEDULA and "ejemplo" in ref_a5:
        return SKIP_ROWS + 1   # saltar la fila ejemplo legacy: datos desde row 6
    return SKIP_ROWS           # datos desde row 5


def _build_absence_name_map(wb) -> Dict[str, str]:
    """
    Lee la hoja 'Ausencias' (index 2) para construir un mapa nombre→código.
    Funciona con la plantilla nueva (display=name: col A=nombre, col B=código)
    y con plantillas legacy (col A=código, col B=nombre) — en ese caso el mapa queda vacío
    porque los valores en col A ya son códigos y no necesitan traducción.
    """
    try:
        ws = wb.worksheets[2]  # Reporte(0), Turnos(1), Ausencias(2)
    except IndexError:
        return {}
    result: Dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = _str(row[0]) if row and len(row) > 0 else ""
        code = _str(row[1]) if row and len(row) > 1 else ""
        if name and code and not name.isdigit():
            result[name] = code
    return result


def parse_import_file(file_bytes: bytes) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(filename=io.BytesIO(file_bytes), data_only=True)

    absence_name_to_code = _build_absence_name_map(wb)

    # Use first sheet regardless of name
    ws = wb.worksheets[0]
    skip = _detect_skip_rows(ws)
    rows_out = []

    for raw_row_idx, row in enumerate(ws.iter_rows(min_row=skip + 1, values_only=True)):
        # Skip fully empty rows
        if all(v is None or _str(v) == "" for v in row):
            continue

        row_number = raw_row_idx + skip + 1
        errors: List[str] = []

        identificacion = _str(row[COL_IDENTIFICACION] if len(row) > COL_IDENTIFICACION else None)
        turno          = _extract_code(row[COL_TURNO]         if len(row) > COL_TURNO         else None) or None
        tipo_ausencia_raw = _str(row[COL_TIPO_AUSENCIA] if len(row) > COL_TIPO_AUSENCIA else None)
        if tipo_ausencia_raw:
            tipo_ausencia = absence_name_to_code.get(tipo_ausencia_raw) or _extract_code(tipo_ausencia_raw) or None
        else:
            tipo_ausencia = None
        es_ausencia    = bool(tipo_ausencia)
        tipo_bono      = _extract_code(row[COL_TIPO_BONO]     if len(row) > COL_TIPO_BONO     else None) or None
        proyecto       = _str(row[COL_PROYECTO]               if len(row) > COL_PROYECTO       else None) or None
        actividad      = _str(row[COL_ACTIVIDAD]              if len(row) > COL_ACTIVIDAD      else None) or None
        hora_ingreso   = _time_field(row[COL_HORA_INGRESO] if len(row) > COL_HORA_INGRESO else None)
        hora_salida    = _time_field(row[COL_HORA_SALIDA]  if len(row) > COL_HORA_SALIDA  else None)
        salida_dia_sig = _bool_field(row[COL_SALIDA_DIA_SIG] if len(row) > COL_SALIDA_DIA_SIG else None)
        notas          = _str(row[COL_NOTAS]           if len(row) > COL_NOTAS           else None) or None

        if not identificacion:
            errors.append("Identificación es obligatoria")

        if not turno and not tipo_ausencia:
            errors.append("Debe indicar un Turno o un Tipo de Ausencia")

        if turno and tipo_ausencia:
            errors.append("Opciones excluyentes: ingresa Turno o Tipo Ausencia, no ambos")

        rows_out.append({
            "rowIndex":       row_number,
            "identificacion": identificacion,
            "esAusencia":     es_ausencia,
            "turno":          turno,
            "tipoAusencia":   tipo_ausencia,
            "tipoBono":       tipo_bono,
            "proyecto":       proyecto,
            "actividad":      actividad,
            "horaIngreso":    hora_ingreso,
            "horaSalida":     hora_salida,
            "salidaDiaSiguiente": salida_dia_sig,
            "notas":          notas,
            "parseErrors":    errors,
        })

    return rows_out
