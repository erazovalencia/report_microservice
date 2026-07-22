from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from typing import List

from ...base import BaseExportService
from ....models.AttendanceModel import AttendanceHistoryExportRow

COLORS = {
    "header_bg": "1F3864",
    "header_fg": "FFFFFF",
    "row_even":  "F5F5F5",
    "row_odd":   "FFFFFF",
    "in_bg":     "D9E8D9",
    "out_bg":    "FDEBD0",
}

TIPO_LABEL = {
    "IN":     "Entrada",
    "OUT":    "Salida",
    "ACCESS": "Acceso",
    "OTHER":  "Otro",
}

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

HEADERS = [
    ("Cédula",      14),
    ("Nombre",      34),
    ("Fecha",       14),
    ("Hora",        10),
    ("Tipo",        14),
    ("Dispositivo", 20),
]


def _row_bg(tipo: str, row_index: int) -> str:
    if tipo == "IN":
        return COLORS["in_bg"] if row_index % 2 == 0 else "EAF4EA"
    if tipo == "OUT":
        return COLORS["out_bg"] if row_index % 2 == 0 else "FEF5E7"
    return COLORS["row_even"] if row_index % 2 == 0 else COLORS["row_odd"]


class AttendanceHistoryExportService(BaseExportService):

    def generate_file(self, data: List[AttendanceHistoryExportRow], options=None) -> io.BytesIO:
        wb = Workbook()
        self._build_sheet(wb, data)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def _build_sheet(self, wb: Workbook, rows: List[AttendanceHistoryExportRow]):
        ws = wb.active
        ws.title = "Histórico Asistencia"
        ws.freeze_panes = "A3"

        ws.merge_cells(f"A1:{get_column_letter(len(HEADERS))}1")
        title_cell = ws["A1"]
        title_cell.value = "HISTÓRICO DE ASISTENCIA — Erazo Valencia"
        title_cell.font = Font(bold=True, size=13, color=COLORS["header_fg"])
        title_cell.fill = PatternFill(start_color=COLORS["header_bg"], fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        for col, (label, width) in enumerate(HEADERS, start=1):
            cell = ws.cell(row=2, column=col, value=label)
            cell.font = Font(bold=True, size=10, color=COLORS["header_fg"])
            cell.fill = PatternFill(start_color=COLORS["header_bg"], fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[2].height = 24

        for i, row in enumerate(rows):
            r = i + 3
            bg = _row_bg(row.tipo, i)
            fill = PatternFill(start_color=bg, fill_type="solid")

            values = [
                row.cedula,
                row.nombre,
                row.fecha,
                row.hora,
                TIPO_LABEL.get(row.tipo, row.tipo),
                row.dispositivo or "",
            ]

            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.fill = fill
                cell.border = BORDER
                cell.alignment = Alignment(vertical="center", horizontal="center" if col in (1, 3, 4, 5) else "left")
                cell.font = Font(size=9)

            ws.row_dimensions[r].height = 16

        ws.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS))}{len(rows) + 2}"

    def get_content_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def get_file_extension(self) -> str:
        return ".xlsx"
