from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from typing import List
from datetime import datetime, timezone, timedelta

BOGOTA_TZ = timezone(timedelta(hours=-5))

from ...base import BaseExportService
from ....models.VacationsModel import VacationExportRow

COLORS = {
    "header_bg": "1F3864",
    "header_fg": "FFFFFF",
}

ESTADO_LABEL = {
    "pending":     "Pendiente",
    "approved":    "Aprobada",
    "in_progress": "En curso",
    "taken":       "Tomada",
    "rejected":    "Rechazada",
    "cancelled":   "Cancelada",
}

ESTADO_BG = {
    "pending":     "FEF3C7",
    "approved":    "DBEAFE",
    "in_progress": "FDEBD0",
    "taken":       "DCFCE7",
    "rejected":    "FEE2E2",
    "cancelled":   "E2E8F0",
}

TIPO_LABEL = {
    "TIME":  "Tiempo",
    "MONEY": "Dinero",
    "BOTH":  "Mixta",
}

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

HEADERS = [
    ("Fecha Solicitud",  18),
    ("Cédula",           14),
    ("Empleado",         32),
    ("Tipo",             10),
    ("Inicio",           12),
    ("Fin",              12),
    ("Estado",           13),
    ("Saldo",            10),
    ("Días Tiempo",      12),
    ("Días Dinero",      12),
    ("Disponible",       11),
    ("Aprobador",        28),
    ("Fecha Decisión",   18),
    ("Días Tomados",     13),
    ("Motivo",           35),
    ("Motivo Rechazo",   35),
    ("Observaciones",    35),
]


def _fmt_datetime(iso: str) -> str:
    if not iso:
        return ""
    try:
        dt = datetime.fromisoformat(iso.replace("Z", "+00:00"))
        return dt.astimezone(BOGOTA_TZ).strftime("%d/%m/%Y %H:%M")
    except Exception:
        return iso


def _fmt_date(iso: str) -> str:
    if not iso:
        return ""
    try:
        return datetime.fromisoformat(iso[:10]).strftime("%d/%m/%Y")
    except Exception:
        return iso


class VacationReportExportService(BaseExportService):

    def generate_file(self, data: List[VacationExportRow], options=None) -> io.BytesIO:
        wb = Workbook()
        self._build_report_sheet(wb, data)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def _build_report_sheet(self, wb: Workbook, rows: List[VacationExportRow]):
        ws = wb.active
        ws.title = "Vacaciones"
        ws.freeze_panes = "A3"

        # Title row
        ws.merge_cells(f"A1:{get_column_letter(len(HEADERS))}1")
        title_cell = ws["A1"]
        title_cell.value = "SOLICITUDES DE VACACIONES — Erazo Valencia"
        title_cell.font = Font(bold=True, size=13, color=COLORS["header_fg"])
        title_cell.fill = PatternFill(start_color=COLORS["header_bg"], fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        # Header row
        for col, (label, width) in enumerate(HEADERS, start=1):
            cell = ws.cell(row=2, column=col, value=label)
            cell.font = Font(bold=True, size=10, color=COLORS["header_fg"])
            cell.fill = PatternFill(start_color=COLORS["header_bg"], fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[2].height = 28

        # Data rows
        for i, row in enumerate(rows):
            r = i + 3
            bg = ESTADO_BG.get(row.estado, "FFFFFF" if i % 2 else "F5F5F5")
            fill = PatternFill(start_color=bg, fill_type="solid")

            values = [
                _fmt_datetime(row.fechaSolicitud),
                row.identificacion,
                row.nombre,
                TIPO_LABEL.get(row.tipo, row.tipo),
                _fmt_date(row.inicio),
                _fmt_date(row.fin),
                ESTADO_LABEL.get(row.estado, row.estado),
                row.saldo,
                row.diasTiempo if row.diasTiempo else "",
                row.diasDinero if row.diasDinero else "",
                row.disponible,
                row.aprobador,
                _fmt_datetime(row.fechaDecision),
                row.diasTomados if row.diasTomados else "",
                row.motivo,
                row.motivoRechazo,
                row.observaciones,
            ]

            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.fill = fill
                cell.border = BORDER
                # cols centradas: fechas, cédula, tipo, estado y numéricas
                cell.alignment = Alignment(
                    vertical="center",
                    horizontal="center" if col in (1, 2, 4, 5, 6, 7, 8, 9, 10, 11, 13, 14) else "left",
                )
                cell.font = Font(size=9)

            ws.row_dimensions[r].height = 16

        # Auto-filter on header row
        ws.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS))}{len(rows) + 2}"

    def get_content_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def get_file_extension(self) -> str:
        return ".xlsx"
