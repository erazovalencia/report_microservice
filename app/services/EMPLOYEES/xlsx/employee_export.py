from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from typing import List

from ...base import BaseExportService
from ....models.EmployeeModel import EmployeeExportRequest

COLORS = {
    "header_bg": "1F3864",
    "header_fg": "FFFFFF",
    "row_even":  "F5F5F5",
    "row_odd":   "FFFFFF",
    "inactive":  "FDEBD0",
}

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


class EmployeeExportService(BaseExportService):

    def generate_file(self, data: EmployeeExportRequest, options=None) -> io.BytesIO:
        wb = Workbook()
        self._build_sheet(wb, data)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def _build_sheet(self, wb: Workbook, payload: EmployeeExportRequest):
        ws = wb.active
        ws.title = "Empleados"
        ws.freeze_panes = "A3"

        columns = payload.columns
        col_count = len(columns)

        ws.merge_cells(f"A1:{get_column_letter(col_count)}1")
        title_cell = ws["A1"]
        title_cell.value = f"{payload.title} — Erazo Valencia"
        title_cell.font = Font(bold=True, size=13, color=COLORS["header_fg"])
        title_cell.fill = PatternFill(start_color=COLORS["header_bg"], fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        for col, column in enumerate(columns, start=1):
            cell = ws.cell(row=2, column=col, value=column.label)
            cell.font = Font(bold=True, size=10, color=COLORS["header_fg"])
            cell.fill = PatternFill(start_color=COLORS["header_bg"], fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER
            ws.column_dimensions[get_column_letter(col)].width = column.width
        ws.row_dimensions[2].height = 24

        for i, row in enumerate(payload.rows):
            r = i + 3
            is_inactive = row.get("estado") == "Retirado"
            bg = COLORS["inactive"] if is_inactive else (COLORS["row_even"] if i % 2 == 0 else COLORS["row_odd"])
            fill = PatternFill(start_color=bg, fill_type="solid")

            for col, column in enumerate(columns, start=1):
                cell = ws.cell(row=r, column=col, value=row.get(column.key, ""))
                cell.fill = fill
                cell.border = BORDER
                cell.alignment = Alignment(vertical="center", horizontal="left")
                cell.font = Font(size=9)

            ws.row_dimensions[r].height = 16

        ws.auto_filter.ref = f"A2:{get_column_letter(col_count)}{len(payload.rows) + 2}"

    def get_content_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def get_file_extension(self) -> str:
        return ".xlsx"
